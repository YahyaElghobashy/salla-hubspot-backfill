#!/usr/bin/env python3
"""Zid export -> the canonical Salla order dict the engine already consumes.

The single most important decision in this module: the output shape is NOT a
new schema. It is exactly the nested dict `Engine.process_order` reads today,
because the entire property map, catalog gate and item router are written
against that shape. Inventing a parallel schema would mean reimplementing all
of it.

Second most important: every item is emitted with `"product": None`. Zid line
items carry a SKU and no Salla product id, which is precisely the shape the
v2.5 legacy-SKU path was built for ("a product deleted in Salla: no id, only a
SKU"). The engine therefore needs no change to handle Zid items in the
standalone case.

Two things must never leak, because both would poison existing lookups:
  * the rich file's `product_id` is a Zid GUID, NOT a Salla product id. If it
    reached `item.product.id` the gate would search `salla_product_id EQ
    <guid>`, miss forever, and hold every order.
  * the rich file's `customer_id` is a Zid id. `create_contact` writes
    `salla_customer_id` from `customer.id`, so a Zid value there would corrupt
    `search_contact_retry` and the customer_sync ledger. It travels as
    side-channel data instead.

Two source formats collapse into that one shape:
  LEGACY (2020..2025, 38 cols) line-item rows, Arabic statuses, order id in `id`
  RICH   (misnamed 2026.xlsx, 69 cols) line-item rows, English statuses,
         product ids, clean SKUs, spans 2025-07-01..2026-02-23

Where both cover a date the RICH file wins wholesale (operator decision
2026-08-10): better fields, English statuses that map straight to pipeline
stages, and it sidesteps the legacy column-swap defect.
"""

import gzip
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

TOKEN_RE = re.compile(r"[A-Z]{1,3}\d{1,3}")
SKU_SHAPE_RE = re.compile(r"^(?:[A-Z]{1,3}\d{1,3})+$", re.I)
BARCODE_RE = re.compile(r"^\d{8,14}$")

RIYADH = "Asia/Riyadh"
SA_CC = "SA"

# ---------------------------------------------------------------------------
# phone
# ---------------------------------------------------------------------------


def zed_phone_key(raw):
    """Normalise to E.164-ish '+966XXXXXXXXX'. Returns "" when unusable.

    The engine keys contacts on (mobile_code, mobile) and searches `phone EQ
    mobile OR phone EQ code+mobile`; engine-created contacts store '966...'
    with no plus. The ~817k pre-existing non-Salla contacts have unknown
    formatting, so the matcher indexes every variant -- this function only has
    to be deterministic, not to guess the portal's style.
    """
    s = str(raw or "").strip()
    if not s:
        return ""
    plus = s.startswith("+")
    d = re.sub(r"\D", "", s)
    if not d:
        return ""
    if d.startswith("00"):
        d = d[2:]
    if len(d) == 12 and d.startswith("966"):
        return "+" + d
    if len(d) == 10 and d.startswith("05"):
        return "+966" + d[1:]
    if len(d) == 9 and d.startswith("5"):
        return "+966" + d
    if plus and 10 <= len(d) <= 15:
        return "+" + d
    if 10 <= len(d) <= 15:
        return "+" + d
    return ""


def split_phone(e164):
    """'+966501234567' -> ('966', '501234567') for the engine's two fields."""
    if not e164 or not e164.startswith("+"):
        return "", ""
    d = e164[1:]
    return (d[:3], d[3:]) if d.startswith("966") else (d[:3], d[3:])


# ---------------------------------------------------------------------------
# sku
# ---------------------------------------------------------------------------


def classify_sku(raw):
    """('single'|'composite'|'barcode'|'malformed'|'empty', tokens)."""
    s = str(raw or "").strip()
    if not s:
        return "empty", []
    if BARCODE_RE.match(s):
        return "barcode", [s]
    up = s.upper()
    toks = TOKEN_RE.findall(up)
    if toks and "".join(toks) == up:
        return ("composite" if len(toks) > 1 else "single"), toks
    return "malformed", []


def is_composite(raw, singles=None):
    """True when the SKU is two or more tokens AND (when a corpus-wide set of
    single-token SKUs is supplied) every token exists on its own. The extra
    check is what stops a barcode or a typo being read as a bundle."""
    kind, toks = classify_sku(raw)
    if kind != "composite":
        return False
    return True if singles is None else all(t in singles for t in toks)


# ---------------------------------------------------------------------------
# status
# ---------------------------------------------------------------------------

# Arabic (legacy) and English (rich) order statuses -> Salla slugs, so the
# existing config.status_stage_map drives the pipeline stage. Authored from
# the census, not guessed; anything unseen raises rather than defaulting
# silently, because a wrong stage on ~974k orders is expensive to undo.
STATUS_AR = {
    "تم التوصيل": ("delivered", "Delivered"),
    "تم الإلغاء": ("canceled", "Canceled"),
    "ملغي": ("canceled", "Canceled"),
    "جاري التوصيل": ("delivering", "Delivering"),
    "قيد التوصيل": ("delivering", "Delivering"),
    "جاهز": ("in_progress", "Ready"),
    "جديد": ("in_progress", "New"),
    "مسترجع": ("restored", "Restored"),
    "مسترجع جزئي": ("restored", "Partially restored"),
    "تجهيز": ("in_progress", "Preparing"),
    "قيد التنفيذ": ("in_progress", "In progress"),
    "بانتظار الدفع": ("payment_pending", "Payment pending"),
    "قيد المراجعة": ("under_review", "Under review"),
    "تم التنفيذ": ("completed", "Completed"),
}
STATUS_EN = {
    "delivered": ("delivered", "Delivered"),
    "canceled": ("canceled", "Canceled"),
    "cancelled": ("canceled", "Canceled"),
    "in delivery": ("delivering", "Delivering"),
    "delivering": ("delivering", "Delivering"),
    "new": ("in_progress", "New"),
    "ready": ("in_progress", "Ready"),
    "prepairing": ("in_progress", "Preparing"),
    "preparing": ("in_progress", "Preparing"),
    "returned": ("restored", "Restored"),
    "completed": ("completed", "Completed"),
    "payment pending": ("payment_pending", "Payment pending"),
    "under review": ("under_review", "Under review"),
}


class UnmappedStatus(KeyError):
    """Raised rather than defaulting: an unseen status must be a decision."""


def status_slug(raw):
    s = str(raw or "").strip()
    if not s or s.lower() in ("none", "null"):
        return "", ""
    if s in STATUS_AR:
        return STATUS_AR[s]
    low = s.lower()
    if low in STATUS_EN:
        return STATUS_EN[low]
    raise UnmappedStatus(s)


# ---------------------------------------------------------------------------
# row helpers
# ---------------------------------------------------------------------------


def _f(v):
    """Money/number cell -> str the engine can pass through. '' when absent."""
    if v is None or v == "":
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v).strip()
    return str(int(f)) if f == int(f) else str(f)


def _dt(v):
    """Cell -> 'YYYY-MM-DD HH:MM:SS'. The legacy columns are labelled
    (Asia/Riyadh) and the rich order_date is naive local, so the timezone is
    stamped literally by the mapper -- a naive/UTC pass would shift every
    hs_external_created_date by three hours."""
    if v is None:
        return ""
    s = str(v).strip()
    # The 12-hour form MUST be tried first. The legacy files write
    # "2020-06-29 07:03 PM"; a generic HH:MM match would happily take "07:03"
    # and silently drop the PM, shifting half of six years of orders by 12
    # hours. Regression covered by TestTimestamps.test_legacy_12h_clock.
    m = re.match(r"^(\d{4}-\d{2}-\d{2})\s+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*([AP])\.?M\.?$",
                 s, re.I)
    if m:
        h = int(m.group(2)) % 12 + (12 if m.group(5).upper() == "P" else 0)
        return f"{m.group(1)} {h:02d}:{m.group(3)}:{m.group(4) or '00'}"
    m = re.match(r"^(\d{4}-\d{2}-\d{2})[ T](\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        return (f"{m.group(1)} {int(m.group(2)):02d}:{m.group(3)}:"
                f"{m.group(4) or '00'}")
    m = re.match(r"^(\d{4}-\d{2}-\d{2})$", s)
    return f"{m.group(1)} 00:00:00" if m else s


def split_name(full):
    s = str(full or "").strip()
    if not s:
        return "", ""
    parts = s.split(None, 1)
    return parts[0], (parts[1] if len(parts) > 1 else "")


def is_junk_row(status, mobile):
    """A row with neither a status nor a phone carries nothing usable. Counted
    and logged rather than silently skipped, because the reconciliation chain
    (source -> kept -> orders -> planned -> held -> emitted) must sum."""
    return (status is None or str(status).strip() == "") and not str(mobile or "").strip()


def repair_column_swap(sku, name):
    """Legacy 2025 defect: `sku` holds a small integer (really the quantity)
    and `product name` holds the SKU. Returns (sku, name, qty_hint, repaired).

    Left unrepaired this writes line items with hs_sku="1", which is why any
    post-repair SKU that still fails shape validation fails the whole month.
    """
    s, n = str(sku or "").strip(), str(name or "").strip()
    if s.isdigit() and len(s) <= 3 and (SKU_SHAPE_RE.match(n)
                                        or BARCODE_RE.match(n)):
        # Two variants, both real in the 2025 file: the displaced value is a
        # C-format SKU ("C18") or a barcode ("6287032431307"). The barcode
        # variant accounts for ~10,900 rows on its own; missing it would write
        # line items with hs_sku="1".
        return n, "", s, True
    return s, n, "", False


# ---------------------------------------------------------------------------
# mappers
# ---------------------------------------------------------------------------


class Mapper:
    """Shared assembly of the canonical order dict from grouped line rows."""

    format = "base"

    def order_key(self, row, ix):
        raise NotImplementedError

    def row_to_item(self, row, ix, seq, oid, names_by_sku):
        raise NotImplementedError

    def order_head(self, row, ix, oid, items):
        raise NotImplementedError

    def build(self, oid, rows, ix, names_by_sku):
        """Group line rows into one canonical order dict, or None when the
        order retains no usable item."""
        items = []
        for seq, r in enumerate(rows, 1):
            it = self.row_to_item(r, ix, seq, oid, names_by_sku)
            if it is not None:
                items.append(it)
        if not items:
            return None
        order = self.order_head(rows[0], ix, oid, items)
        order["items"] = items
        return order


class LegacyMapper(Mapper):
    format = "legacy"

    def order_key(self, row, ix):
        v = row[ix["id"]]
        return str(v) if v is not None else ""

    def row_to_item(self, row, ix, seq, oid, names_by_sku):
        sku, name, qty_hint, repaired = repair_column_swap(
            row[ix["sku"]], row[ix["product name"]])
        if not sku:
            return None
        if not name:
            name = names_by_sku.get(sku, "") or sku
        qty = _f(row[ix["quantity"]]) or qty_hint or "1"
        return {
            # deterministic id: re-running the emitter must not mint new items
            "id": f"Z{oid}-{seq}",
            "name": name,
            "sku": sku,
            "quantity": qty,
            "currency": str(row[ix["currency"]] or "SAR"),
            "product_type": "",
            "product": None,          # forces the legacy-SKU path in the engine
            "amounts": {
                "price_without_tax": {"amount": _f(row[ix["unit_price"]])},
                "original_price": {"amount": _f(row[ix["unit_price"]])},
            },
            "_zed": {"repaired_swap": repaired},
        }

    def order_head(self, row, ix, oid, items):
        slug, sname = status_slug(row[ix["order_status"]])
        e164 = zed_phone_key(row[ix["customer_mobile"]])
        code, mob = split_phone(e164)
        first, last = split_name(row[ix["customer_name"]])
        cur = str(row[ix["currency"]] or "SAR")
        coupon = row[ix.get("coupon_code")] if ix.get("coupon_code") is not None else None
        return {
            "id": oid,
            "reference_id": oid,
            "date": {"date": _dt(row[ix["added_at (Asia/Riyadh)"]]),
                     "timezone": RIYADH},
            "status": {"slug": slug, "name": sname},
            "payment_method": str(row[ix["payment_method"]] or ""),
            "urls": {"admin": ""},
            "customer": {
                "id": "",                      # never a Zid id: see module docstring
                "first_name": first, "last_name": last,
                "full_name": str(row[ix["customer_name"]] or ""),
                "email": str(row[ix["customer_email"]] or ""),
                "mobile_code": code, "mobile": mob,
                "city": str(row[ix["shipping_city"]] or ""),
                "country": "", "country_code": SA_CC,
                "urls": {"admin": ""},
                "created_at": {"date": ""},
            },
            "amounts": {
                "total": {"amount": _f(row[ix["total"]])},
                "sub_total": {"amount": _f(row[ix["sub_totals"]])},
                "shipping_cost": {"amount": _f(row[ix["shipping"]]),
                                  "currency": cur},
                "tax": {"amount": {"amount": _f(row[ix["vat"]])}},
                "discounts": ([{"code": str(coupon)}] if coupon else []),
            },
            "_zed": {"source_file": "legacy", "phone_e164": e164},
        }


class RichMapper(Mapper):
    format = "rich"

    def order_key(self, row, ix):
        v = row[ix["order_id"]]
        return str(v) if v is not None else ""

    def row_to_item(self, row, ix, seq, oid, names_by_sku):
        sku = str(row[ix["product_sku"]] or "").strip()
        if not sku:
            return None
        qcol = ix.get("Quantity")
        qty = _f(row[qcol]) if qcol is not None else "1"
        return {
            "id": f"Z{oid}-{seq}",
            "name": str(row[ix["product_name"]] or sku),
            "sku": sku,
            "quantity": qty or "1",
            "currency": str(row[ix["order_currency_code"]] or "SAR"),
            "product_type": "",
            "product": None,          # Zid product_id is NOT a Salla id
            "amounts": {
                "price_without_tax": {"amount": _f(row[ix["net_price"]])},
                "original_price": {"amount": _f(row[ix["gross_price"]])},
            },
            "_zed": {"zid_product_id": str(row[ix["product_id"]] or ""),
                     "name_ar": str(row[ix.get("product_name_ar")] or "")
                     if ix.get("product_name_ar") is not None else ""},
        }

    def order_head(self, row, ix, oid, items):
        slug, sname = status_slug(row[ix["order_status_name"]])
        e164 = zed_phone_key(row[ix["customer_telephone"]])
        code, mob = split_phone(e164)
        first, last = split_name(row[ix["customer_name"]])
        cur = str(row[ix["order_currency_code"]] or "SAR")
        return {
            "id": oid,
            "reference_id": str(row[ix["order_code"]] or oid),
            "date": {"date": _dt(row[ix["order_date"]]), "timezone": RIYADH},
            "status": {"slug": slug, "name": sname},
            "payment_method": str(row[ix["payment_method_name"]] or ""),
            "urls": {"admin": ""},
            "customer": {
                "id": "",
                "first_name": first, "last_name": last,
                "full_name": str(row[ix["customer_name"]] or ""),
                "email": str(row[ix["customer_email"]] or ""),
                "mobile_code": code, "mobile": mob,
                "city": str(row[ix["city_name"]] or ""),
                "country": "", "country_code": SA_CC,
                "urls": {"admin": ""},
                "created_at": {"date": ""},
            },
            "amounts": {
                "total": {"amount": _f(row[ix["total_value"]])},
                "sub_total": {"amount": _f(row[ix["sub_total_value"]])},
                "shipping_cost": {"amount": _f(row[ix["shipping_fees"]]),
                                  "currency": cur},
                "tax": {"amount": {"amount": _f(row[ix["vat_value"]])}},
                "discounts": [],
            },
            "_zed": {"source_file": "rich", "phone_e164": e164,
                     "zid_customer_id": str(row[ix["customer_id"]] or ""),
                     "tracking_id": str(row[ix.get("order_tracking_id")] or "")
                     if ix.get("order_tracking_id") is not None else ""},
        }


def mapper_for(ix):
    if "order_id" in ix:
        return RichMapper()
    if "id" in ix:
        return LegacyMapper()
    raise ValueError("unrecognised sheet: neither order_id nor id present")


# ---------------------------------------------------------------------------
# driver
# ---------------------------------------------------------------------------


def _sheet(z, member):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(z.read(member)), read_only=True,
                               data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [h for h in next(it)]
    return wb, it, {h: i for i, h in enumerate(hdr) if h is not None}


def dominant_names(z, members):
    """SKU -> most common product name across the whole corpus.

    Column-swapped rows lose their product name (the SKU was sitting in it),
    so the name has to come from the rest of the corpus rather than the row.
    """
    names = defaultdict(Counter)
    for m in members:
        wb, it, ix = _sheet(z, m)
        sku_c = ix.get("sku", ix.get("product_sku"))
        nm_c = ix.get("product name", ix.get("product_name"))
        if sku_c is None or nm_c is None:
            wb.close()
            continue
        for r in it:
            s, n = r[sku_c], r[nm_c]
            if s and n:
                s2, n2, _, _ = repair_column_swap(s, n)
                if s2 and n2:
                    names[s2][n2.strip()[:80]] += 1
        wb.close()
    return {s: c.most_common(1)[0][0] for s, c in names.items()}


def normalize(zip_path, outdir="mirror/zed", repairs_log="mirror/zed_repairs.csv",
              dropped_log="mirror/zed_dropped.csv"):
    """Stream every workbook into one gzipped JSONL per calendar month.

    The rich file wins wherever it covers a date (operator decision): its ids
    are collected first, and any legacy row whose order id is already claimed
    is skipped rather than merged, so the 173,756-order Jul-Dec 2025 overlap
    cannot double.
    """
    import openpyxl  # noqa: F401  (imported for the error if it's missing)
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    z = zipfile.ZipFile(zip_path)
    members = sorted(n for n in z.namelist()
                     if n.endswith(".xlsx") and "__MACOSX" not in n)
    # rich first so its ids claim the overlap
    members.sort(key=lambda m: 0 if "2026" in m else 1)

    names_by_sku = dominant_names(z, members)
    claimed = set()
    by_month = defaultdict(list)
    stats = {"source_rows": 0, "junk_rows": 0, "repaired_rows": 0,
             "overlap_skipped_orders": 0, "orders": 0, "items": 0,
             "orders_dropped_no_items": 0, "unmapped_status": Counter()}
    repairs, dropped = [], []

    for m in members:
        label = m.split("/")[-1].replace(".xlsx", "")
        wb, it, ix = _sheet(z, m)
        mapper = mapper_for(ix)
        oid_getter = (lambda r: r[ix["order_id"]]) if mapper.format == "rich" \
            else (lambda r: r[ix["id"]])
        st_c = ix.get("order_status_name", ix.get("order_status"))
        mob_c = ix.get("customer_telephone", ix.get("customer_mobile"))
        dt_c = ix.get("order_date", ix.get("added_at (Asia/Riyadh)"))

        groups = defaultdict(list)
        for r in it:
            if oid_getter(r) is None and (dt_c is None or r[dt_c] is None):
                continue                      # trailing blank padding
            stats["source_rows"] += 1
            if is_junk_row(r[st_c] if st_c is not None else None,
                           r[mob_c] if mob_c is not None else None):
                stats["junk_rows"] += 1
                dropped.append((label, str(oid_getter(r) or ""), "junk row"))
                continue
            oid = str(oid_getter(r) or "")
            if not oid:
                continue
            groups[oid].append(r)
        wb.close()

        for oid, rows in groups.items():
            if oid in claimed:
                stats["overlap_skipped_orders"] += 1
                continue
            try:
                order = mapper.build(oid, rows, ix, names_by_sku)
            except UnmappedStatus as e:
                stats["unmapped_status"][str(e)] += 1
                dropped.append((label, oid, f"unmapped status {e}"))
                continue
            if order is None:
                stats["orders_dropped_no_items"] += 1
                dropped.append((label, oid, "no usable items"))
                continue
            claimed.add(oid)
            for i in order["items"]:
                if i.get("_zed", {}).get("repaired_swap"):
                    stats["repaired_rows"] += 1
                    repairs.append((label, oid, i["sku"], i["quantity"]))
            month = (order["date"]["date"] or "0000-00")[:7]
            by_month[month].append(order)
            stats["orders"] += 1
            stats["items"] += len(order["items"])
        print(f"  {label}: {stats['orders']:,} orders so far", flush=True)

    for month, orders in sorted(by_month.items()):
        with gzip.open(out / f"{month}.jsonl.gz", "wt", encoding="utf-8") as f:
            for o in orders:
                f.write(json.dumps(o, ensure_ascii=False) + "\n")

    Path(repairs_log).parent.mkdir(parents=True, exist_ok=True)
    with open(repairs_log, "w", encoding="utf-8") as f:
        f.write("file,order_id,repaired_sku,quantity\n")
        for row in repairs:
            f.write(",".join(str(x) for x in row) + "\n")
    with open(dropped_log, "w", encoding="utf-8") as f:
        f.write("file,order_id,reason\n")
        for row in dropped:
            f.write(",".join(str(x).replace(",", " ") for x in row) + "\n")

    stats["unmapped_status"] = dict(stats["unmapped_status"])
    stats["months"] = {m: len(v) for m, v in sorted(by_month.items())}
    return stats


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", required=True)
    ap.add_argument("--out", default="mirror/zed")
    args = ap.parse_args()
    s = normalize(args.zip, args.out)
    print("\n=========== NORMALISE ===========")
    print(f"source rows          {s['source_rows']:,}")
    print(f"  junk dropped       {s['junk_rows']:,}")
    print(f"  column-swaps fixed {s['repaired_rows']:,}")
    print(f"overlap skipped      {s['overlap_skipped_orders']:,} orders "
          f"(already claimed by the rich file)")
    print(f"orders written       {s['orders']:,}")
    print(f"items written        {s['items']:,}")
    print(f"dropped, no items    {s['orders_dropped_no_items']:,}")
    if s["unmapped_status"]:
        print(f"UNMAPPED STATUSES    {s['unmapped_status']}")
    print(f"months               {len(s['months'])}")
    print(f"-> {args.out}")


if __name__ == "__main__":
    main()
