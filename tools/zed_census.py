#!/usr/bin/env python3
"""Phase 0 census over the Zid export. Read-only, no HubSpot, no network.

Answers the questions the import plan is gated on, before any code is written
against assumptions:

  1. STATUS COVERAGE   every distinct order_status / order_status_name value,
                       with counts, so the slug map can be authored from the
                       data instead of guessed. 100% coverage is required
                       before any month emits.
  2. SKU UNIVERSE      distinct SKUs, split into single-token / composite /
                       barcode / malformed, with the dominant product name and
                       its consistency share per SKU. This is the seed for the
                       monthly approval sheets.
  3. DEFECTS           the legacy column-swap (sku numeric, product name looks
                       like a SKU) and junk rows (null status + null mobile),
                       counted per file so the reconciliation chain can be
                       proven later.
  4. PHONE SHAPES      how customer phone numbers are actually formatted, which
                       decides the normaliser's rules.
  5. RECONCILIATION    source rows -> kept rows -> orders, per month, per file,
                       plus the legacy/rich overlap by order id.

Usage:
    python3 tools/zed_census.py --zip "/path/ZID ORDERS.zip" --out mirror/zed_census.json
"""

import argparse
import collections
import io
import json
import re
import sys
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# A SKU token is letters+digits, e.g. C18, CH11, CP1. A composite is two or
# more tokens concatenated with nothing between them (C18CH11CH10).
TOKEN_RE = re.compile(r"[A-Z]{1,3}\d{1,3}")
SKU_SHAPE_RE = re.compile(r"^(?:[A-Z]{1,3}\d{1,3})+$", re.I)
BARCODE_RE = re.compile(r"^\d{8,14}$")

LEGACY_COLS = {"id", "order_status", "customer_mobile", "sku", "product name",
               "added_at (Asia/Riyadh)", "quantity", "unit_price", "total"}
RICH_COLS = {"order_id", "order_status_name", "customer_telephone",
             "product_sku", "product_name", "order_date", "Quantity"}


def classify_sku(raw):
    """single | composite | barcode | malformed | empty"""
    s = str(raw or "").strip()
    if not s:
        return "empty", []
    if BARCODE_RE.match(s):
        return "barcode", [s]
    up = s.upper()
    toks = TOKEN_RE.findall(up)
    if toks and "".join(toks) == up:
        return ("composite" if len(toks) > 1 else "single"), toks
    return "malformed", []


def open_sheet(z, member):
    wb = openpyxl.load_workbook(io.BytesIO(z.read(member)), read_only=True,
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [h for h in next(it)]
    return wb, it, hdr, {h: i for i, h in enumerate(hdr) if h is not None}


def census_file(z, member):
    wb, it, hdr, ix = open_sheet(z, member)
    fmt = "rich" if "order_id" in ix else ("legacy" if "id" in ix else "unknown")
    if fmt == "unknown":
        wb.close()
        return {"format": "unknown", "columns": len(hdr)}

    if fmt == "legacy":
        c_oid, c_st, c_mob = ix["id"], ix["order_status"], ix["customer_mobile"]
        c_sku, c_name = ix["sku"], ix["product name"]
        c_date, c_qty = ix["added_at (Asia/Riyadh)"], ix["quantity"]
        c_email = ix.get("customer_email")
    else:
        c_oid, c_st = ix["order_id"], ix["order_status_name"]
        c_mob = ix["customer_telephone"]
        c_sku, c_name = ix["product_sku"], ix["product_name"]
        c_date, c_qty = ix["order_date"], ix.get("Quantity")
        c_email = ix.get("customer_email")

    rows = kept = junk = swapped = 0
    orders, kept_orders = set(), set()
    statuses = collections.Counter()
    months = collections.Counter()
    sku_kind = collections.Counter()
    sku_names = collections.defaultdict(collections.Counter)
    phone_shape = collections.Counter()
    no_email = 0
    ids_by_month = collections.defaultdict(set)

    for r in it:
        # trailing blank rows (the rich file is padded to Excel's row limit)
        if r[c_oid] is None and r[c_date] is None:
            continue
        rows += 1
        oid = str(r[c_oid]) if r[c_oid] is not None else ""
        if oid:
            orders.add(oid)
        st = r[c_st]
        mob = r[c_mob]
        statuses[str(st) if st is not None else "(null)"] += 1

        if st is None and not mob:
            junk += 1
            continue
        kept += 1
        if oid:
            kept_orders.add(oid)

        d = str(r[c_date] or "")[:10]
        if d:
            months[d[:7]] += 1
            if oid:
                ids_by_month[d[:7]].add(oid)

        sku_raw = r[c_sku]
        name_raw = r[c_name]
        # legacy column-swap: sku holds a small integer, product name holds
        # something SKU-shaped
        s_txt, n_txt = str(sku_raw or "").strip(), str(name_raw or "").strip()
        if (fmt == "legacy" and s_txt.isdigit() and len(s_txt) <= 3
                and SKU_SHAPE_RE.match(n_txt)):
            swapped += 1
            sku_raw, name_raw = name_raw, None

        kind, _ = classify_sku(sku_raw)
        sku_kind[kind] += 1
        if kind != "empty":
            sku_names[str(sku_raw).strip()][str(name_raw or "").strip()[:60]] += 1

        if mob:
            p = re.sub(r"\D", "", str(mob))
            phone_shape[f"len{len(p)}"
                        + ("/966" if p.startswith("966") else
                           "/05" if p.startswith("05") else
                           "/5" if p.startswith("5") else "/other")] += 1
        if c_email is not None and not r[c_email]:
            no_email += 1
    wb.close()

    # dominant name per SKU + consistency share
    sku_profile = {}
    for sku, names in sku_names.items():
        tot = sum(names.values())
        top, n = names.most_common(1)[0]
        kind, toks = classify_sku(sku)
        sku_profile[sku] = {"rows": tot, "name": top,
                            "name_consistency": round(n / tot, 3),
                            "kind": kind, "tokens": toks}

    return {
        "format": fmt, "columns": len(hdr),
        "rows": rows, "kept_rows": kept, "junk_rows": junk,
        "column_swapped_rows": swapped,
        "orders": len(orders), "kept_orders": len(kept_orders),
        "rows_without_email": no_email,
        "statuses": statuses.most_common(),
        "months": dict(sorted(months.items())),
        "sku_kinds": dict(sku_kind),
        "distinct_skus": len(sku_profile),
        "sku_profile": sku_profile,
        "phone_shapes": phone_shape.most_common(12),
        "_ids_by_month": {m: sorted(s) for m, s in ids_by_month.items()},
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", required=True)
    ap.add_argument("--out", default="mirror/zed_census.json")
    args = ap.parse_args()

    z = zipfile.ZipFile(args.zip)
    members = sorted(n for n in z.namelist()
                     if n.endswith(".xlsx") and "__MACOSX" not in n)
    per_file, id_sets = {}, {}
    for m in members:
        label = m.split("/")[-1].replace(".xlsx", "")
        print(f"scanning {label} ...", flush=True)
        c = census_file(z, m)
        ids_by_month = c.pop("_ids_by_month", {})
        id_sets[label] = ids_by_month
        per_file[label] = c
        print(f"  fmt={c['format']} rows={c['rows']:,} kept={c['kept_rows']:,} "
              f"junk={c['junk_rows']:,} swapped={c['column_swapped_rows']:,} "
              f"orders={c['orders']:,} skus={c['distinct_skus']}", flush=True)

    # ---- cross-file overlap by order id -----------------------------------
    rich = [k for k, v in per_file.items() if v["format"] == "rich"]
    legacy = [k for k, v in per_file.items() if v["format"] == "legacy"]
    rich_ids = set()
    for k in rich:
        for m, ids in id_sets[k].items():
            rich_ids.update(ids)
    overlap = {}
    legacy_only_ids = set()
    for k in legacy:
        ids = set()
        for m, mids in id_sets[k].items():
            ids.update(mids)
        inter = ids & rich_ids
        overlap[k] = {"legacy_ids": len(ids), "shared_with_rich": len(inter)}
        legacy_only_ids |= (ids - rich_ids)

    union = len(legacy_only_ids | rich_ids)

    # ---- global SKU universe ----------------------------------------------
    universe = collections.defaultdict(lambda: {"rows": 0, "names": collections.Counter(),
                                                "kind": "", "tokens": [], "files": set()})
    for label, c in per_file.items():
        for sku, p in c.get("sku_profile", {}).items():
            u = universe[sku]
            u["rows"] += p["rows"]
            u["names"][p["name"]] += p["rows"]
            u["kind"] = p["kind"]
            u["tokens"] = p["tokens"]
            u["files"].add(label)
    singles = {s for s, u in universe.items() if u["kind"] == "single"}
    composites = {s: u for s, u in universe.items() if u["kind"] == "composite"}
    resolvable = {s: u for s, u in composites.items()
                  if all(t in singles for t in u["tokens"])}

    out = {
        "per_file": per_file,
        "overlap": overlap,
        "union_orders": union,
        "sku_universe": {
            "distinct": len(universe),
            "single": len(singles),
            "composite": len(composites),
            "composite_fully_resolvable": len(resolvable),
            "barcode": sum(1 for u in universe.values() if u["kind"] == "barcode"),
            "malformed": sum(1 for u in universe.values() if u["kind"] == "malformed"),
            "top": [
                {"sku": s, "rows": u["rows"], "kind": u["kind"],
                 "name": u["names"].most_common(1)[0][0],
                 "tokens": u["tokens"]}
                for s, u in sorted(universe.items(), key=lambda kv: -kv[1]["rows"])[:60]
            ],
        },
    }
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.out).write_text(json.dumps(out, ensure_ascii=False, indent=1))

    print("\n================ CENSUS ================")
    tr = sum(c["rows"] for c in per_file.values())
    tk = sum(c["kept_rows"] for c in per_file.values())
    tj = sum(c["junk_rows"] for c in per_file.values())
    ts = sum(c["column_swapped_rows"] for c in per_file.values())
    print(f"source rows {tr:,} -> kept {tk:,}  (junk dropped {tj:,}, "
          f"column-swaps repaired {ts:,})")
    print(f"union unique orders: {union:,}")
    print("\noverlap by file:")
    for k, v in overlap.items():
        print(f"  {k}: {v['legacy_ids']:,} ids, {v['shared_with_rich']:,} also in rich")
    su = out["sku_universe"]
    print(f"\nSKU universe: {su['distinct']} distinct "
          f"(single {su['single']}, composite {su['composite']} of which "
          f"{su['composite_fully_resolvable']} fully resolvable, "
          f"barcode {su['barcode']}, malformed {su['malformed']})")
    print("\nstatus values needing a slug mapping:")
    allst = collections.Counter()
    for c in per_file.values():
        for s, n in c["statuses"]:
            allst[s] += n
    for s, n in allst.most_common():
        print(f"  {n:>9,}  {s}")
    print(f"\n-> {args.out}")


if __name__ == "__main__":
    main()
