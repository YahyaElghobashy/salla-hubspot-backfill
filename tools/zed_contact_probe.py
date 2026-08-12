#!/usr/bin/env python3
"""Stop-the-line gate: what fraction of Zid customers already exist in HubSpot?

The whole contact plan rests on one premise -- that the ~817,635 HubSpot
contacts with no salla_customer_id ARE the Zid-era customer base. If that is
right, importing ~974k orders costs a few thousand contact creates. If it is
wrong, it costs several hundred thousand, and the plan needs rethinking before
a single write.

This samples Zid phone numbers straight from the export, normalises them the
same way the importer will, and looks them up in the local contact snapshot.
Read-only, no HubSpot calls (the snapshot is already on disk), so it costs
nothing to run and takes about a minute.

Threshold: under 60% the premise is considered broken and the script exits 1.

Usage:
    python3 tools/zed_contact_probe.py --zip "/path/ZID ORDERS.zip" --sample 5000
"""

import argparse
import collections
import io
import json
import random
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import zed_normalize as zn
from zed_snapshot import ContactIndex


def sample_phones(zip_path, per_file, seed=20260810):
    """Reservoir-sample raw phone values from every workbook.

    openpyxl is imported here rather than at module scope: the default path
    samples the normalised months, which needs no Excel reader, and the VM has
    no reason to carry that dependency.
    """
    import openpyxl

    z = zipfile.ZipFile(zip_path)
    members = sorted(n for n in z.namelist()
                     if n.endswith(".xlsx") and "__MACOSX" not in n)
    rnd = random.Random(seed)
    out = []
    for m in members:
        label = m.split("/")[-1].replace(".xlsx", "")
        wb = openpyxl.load_workbook(io.BytesIO(z.read(m)), read_only=True,
                                    data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [h for h in next(it)]
        ix = {h: i for i, h in enumerate(hdr) if h is not None}
        col = ix.get("customer_mobile", ix.get("customer_telephone"))
        dcol = ix.get("added_at (Asia/Riyadh)", ix.get("order_date"))
        if col is None:
            wb.close()
            continue
        res, n = [], 0
        for r in it:
            v = r[col]
            if not v:
                continue
            n += 1
            rec = (label, str(r[dcol])[:7] if dcol is not None else "", str(v))
            if len(res) < per_file:
                res.append(rec)
            else:
                j = rnd.randrange(n)
                if j < per_file:
                    res[j] = rec
        out += res
        wb.close()
        print(f"  sampled {len(res)} from {label} (of {n:,} rows with a phone)",
              flush=True)
    return out


def sample_from_months(months_dir, total, seed=20260810):
    """Sample from the NORMALISED months rather than the raw workbooks.

    Better than sampling the export: these are exactly the orders that will be
    imported, already junk-filtered, overlap-deduped and phone-normalised, so
    the hit rate measured here is the hit rate the import will actually see.
    """
    import glob
    import gzip

    rnd = random.Random(seed)
    files = sorted(glob.glob(str(Path(months_dir) / "*.jsonl.gz")))
    if not files:
        raise SystemExit(f"no normalised months in {months_dir}")
    per_file = max(10, total // len(files))
    out = []
    for fp in files:
        month = Path(fp).name.split(".")[0]
        res, n = [], 0
        with gzip.open(fp, "rt", encoding="utf-8") as f:
            for line in f:
                o = json.loads(line)
                c = o.get("customer") or {}
                raw = f"{c.get('mobile_code','')}{c.get('mobile','')}"
                if not raw.strip():
                    continue
                n += 1
                rec = (month[:4], month, raw)
                if len(res) < per_file:
                    res.append(rec)
                else:
                    j = rnd.randrange(n)
                    if j < per_file:
                        res[j] = rec
        out += res
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--zip", help="sample raw workbooks")
    ap.add_argument("--months", default="mirror/zed",
                    help="sample the normalised months instead (preferred)")
    ap.add_argument("--sample", type=int, default=5000,
                    help="total sample size, split across the sources")
    ap.add_argument("--threshold", type=float, default=60.0,
                    help="percent below which the premise is judged broken")
    args = ap.parse_args()

    idx = ContactIndex()
    print(f"contact snapshot: {idx.count():,} contacts indexed\n")

    if args.zip:
        rows = sample_phones(args.zip, max(50, args.sample // 7))
    else:
        rows = sample_from_months(args.months, args.sample)
        print(f"sampled from normalised months in {args.months}")
    print(f"\ntotal sampled: {len(rows):,}\n")

    hit = miss = unusable = 0
    by_year = collections.defaultdict(lambda: [0, 0])
    miss_examples = []
    for label, month, raw in rows:
        key = zn.zed_phone_key(raw)
        if not key:
            unusable += 1
            continue
        if idx.lookup(key):
            hit += 1
            by_year[label][0] += 1
        else:
            miss += 1
            by_year[label][1] += 1
            if len(miss_examples) < 8:
                miss_examples.append((raw, key, month))

    usable = hit + miss
    pct = (100.0 * hit / usable) if usable else 0.0
    print(f"usable phones   {usable:,}")
    print(f"  matched       {hit:,}  ({pct:.1f}%)")
    print(f"  not found     {miss:,}")
    print(f"unnormalisable  {unusable:,}")
    print("\nby source file:")
    for label in sorted(by_year):
        h, m = by_year[label]
        t = h + m
        print(f"  {label}: {h:,}/{t:,} matched ({100.0*h/t if t else 0:.1f}%)")
    if miss_examples:
        print("\nsample misses (raw -> normalised):")
        for raw, key, month in miss_examples:
            print(f"  {raw!r:>22} -> {key:<16} {month}")

    print()
    if pct >= args.threshold:
        print(f"PASS: {pct:.1f}% >= {args.threshold}% threshold. The existing "
              f"contact base covers most Zid customers; match-or-create is "
              f"viable as planned.")
        return 0
    print(f"FAIL: {pct:.1f}% < {args.threshold}% threshold. The premise that "
          f"the non-Salla HubSpot contacts are the Zid base does NOT hold. "
          f"Contact strategy needs rethinking before any import.")
    return 1


if __name__ == "__main__":
    sys.exit(main())
